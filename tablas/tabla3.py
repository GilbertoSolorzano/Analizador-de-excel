import pandas as pd
from tablas.m_tabla import _make_table 
def tabla_3(df_filtrado: pd.DataFrame, writer, sheet_name: str, col_case, col_qty, col_reason, col_detail_reason, startrow=0):
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # Preparar resumen
    resumen = df_filtrado.loc[:, [col_reason, col_detail_reason, col_case, col_qty]].copy()
    resumen.columns = ['reason', 'detail_reason', 'case_number', 'quantity']
    resumen['quantity'] = pd.to_numeric(resumen['quantity'], errors='coerce').fillna(0)

    # Totales por reason (nivel 1)
    totales_reason = (
        resumen
        .groupby('reason', dropna=False, as_index=False)
        .agg(
            count_of_case_number=('case_number', 'nunique'),
            sum_of_quantity=('quantity', 'sum')
        )
        .sort_values('sum_of_quantity', ascending=False)
    )

    # Detalle por reason + detail_reason (nivel 2)
    detalle = (
        resumen
        .groupby(['reason', 'detail_reason'], dropna=False, as_index=False)
        .agg(
            count_of_case_number=('case_number', 'nunique'),
            sum_of_quantity=('quantity', 'sum')
        )
    )

    # Grand Total
    grand_total_cases = resumen['case_number'].nunique()
    grand_total_qty = resumen['quantity'].sum()

    # Construir tabla jerárquica
    filas = []
    niveles = []
    
    for _, row_reason in totales_reason.iterrows():
        reason = row_reason['reason']
        
        # Nivel 1: Reason
        filas.append({
            'Reason': f"+{reason}",
            'Count of Case Number': int(row_reason['count_of_case_number']),
            'Sum of Quantity': int(row_reason['sum_of_quantity'])
        })
        niveles.append(1)
        
        # Nivel 2: Details de esta reason
        detalles_de_reason = detalle[detalle['reason'] == reason].sort_values('sum_of_quantity', ascending=False)
        for _, det in detalles_de_reason.iterrows():
            filas.append({
                'Reason': f"      • {det['detail_reason']}",
                'Count of Case Number': int(det['count_of_case_number']),
                'Sum of Quantity': int(det['sum_of_quantity'])
            })
            niveles.append(2)
        
        # Línea separadora después de cada reason
        filas.append({
            'Reason': '─' * 50,
            'Count of Case Number': '',
            'Sum of Quantity': ''
        })
        niveles.append(-1)

    # Grand Total
    filas.append({
        'Reason': 'GRAND TOTAL',
        'Count of Case Number': int(grand_total_cases),
        'Sum of Quantity': int(grand_total_qty)
    })
    niveles.append(0)

    tabla_jerarquica = pd.DataFrame(filas)

    # Escribir la tabla
    tabla_jerarquica.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

    # Dar formato
    wb = writer.book
    ws = wb[sheet_name]

    # Estilos
    font_reason = Font(bold=True, size=12, color="1F4E79")
    font_detail = Font(size=10, color="404040")
    font_total = Font(bold=True, size=12, color="FFFFFF")
    font_separator = Font(color="CCCCCC")
    
    fill_reason = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    fill_detail = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_total = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_separator = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
    
    border_bottom = Border(bottom=Side(style='thin', color='CCCCCC'))
    border_thick = Border(bottom=Side(style='medium', color='1F4E79'))

    # Aplicar formato según el nivel
    for i, nivel in enumerate(niveles):
        row_num = startrow + 2 + i
        
        for col in range(1, 4):
            cell = ws.cell(row=row_num, column=col)
            
            if nivel == 1:  # Reason
                cell.font = font_reason
                cell.fill = fill_reason
                cell.border = border_bottom
            elif nivel == 2:  # Detail
                cell.font = font_detail
                cell.fill = fill_detail
            elif nivel == 0:  # Grand Total
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_thick
            elif nivel == -1:  # Separador
                cell.font = font_separator
                cell.fill = fill_separator

    # Formato del encabezado
    for col in range(1, 4):
        cell = ws.cell(row=startrow + 1, column=col)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    print(f"Hoja '{sheet_name}': {len(tabla_jerarquica)} filas en tabla 3.")
    return startrow + len(tabla_jerarquica) + 2