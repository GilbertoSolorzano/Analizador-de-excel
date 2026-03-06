import pandas as pd
from tablas.m_tabla import _make_table 
def tabla_2(df_filtrado: pd.DataFrame, writer, sheet_name: str, col_serie, col_case, col_qty, col_reason, col_detail_reason, startrow=0):
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # Preparar resumen
    resumen = df_filtrado.loc[:, [col_serie, col_reason, col_detail_reason, col_case, col_qty]].copy()
    resumen.columns = ['brand', 'reason', 'detail_reason', 'case_number', 'quantity']
    resumen['quantity'] = pd.to_numeric(resumen['quantity'], errors='coerce').fillna(0)

    # Totales por brand (nivel 1)
    totales_brand = (
        resumen
        .groupby('brand', dropna=False, as_index=False)
        .agg(
            count_of_case_number=('case_number', 'nunique'),
            sum_of_quantity=('quantity', 'sum')
        )
        .sort_values('sum_of_quantity', ascending=False)
    )

    # Totales por brand + reason (nivel 2)
    totales_reason = (
        resumen
        .groupby(['brand', 'reason'], dropna=False, as_index=False)
        .agg(
            count_of_case_number=('case_number', 'nunique'),
            sum_of_quantity=('quantity', 'sum')
        )
    )

    # Detalle por brand + reason + detail_reason (nivel 3)
    detalle = (
        resumen
        .groupby(['brand', 'reason', 'detail_reason'], dropna=False, as_index=False)
        .agg(
            count_of_case_number=('case_number', 'nunique'),
            sum_of_quantity=('quantity', 'sum')
        )
    )

    # Grand Total
    grand_total_cases = resumen['case_number'].nunique()
    grand_total_qty = resumen['quantity'].sum()

    # Construir tabla jerárquica con niveles
    filas = []
    niveles = []  # Para guardar el nivel de cada fila (1=brand, 2=reason, 3=detail, 0=total, -1=separador)
    
    for idx, row_brand in totales_brand.iterrows():
        brand = row_brand['brand']
        
        # Nivel 1: Brand
        filas.append({
            'Brand': f"+ {brand}",
            'Count of Case Number': int(row_brand['count_of_case_number']),
            'Sum of Quantity': int(row_brand['sum_of_quantity'])
        })
        niveles.append(1)
        
        # Nivel 2: Reasons de este brand
        reasons_de_brand = totales_reason[totales_reason['brand'] == brand].sort_values('sum_of_quantity', ascending=False)
        for _, row_reason in reasons_de_brand.iterrows():
            reason = row_reason['reason']
            filas.append({
                'Brand': f"    - {reason}",
                'Count of Case Number': int(row_reason['count_of_case_number']),
                'Sum of Quantity': int(row_reason['sum_of_quantity'])
            })
            niveles.append(2)
            
            # Nivel 3: Details de esta reason
            detalles_de_reason = detalle[(detalle['brand'] == brand) & (detalle['reason'] == reason)].sort_values('sum_of_quantity', ascending=False)
            for _, det in detalles_de_reason.iterrows():
                filas.append({
                    'Brand': f"          • {det['detail_reason']}",
                    'Count of Case Number': int(det['count_of_case_number']),
                    'Sum of Quantity': int(det['sum_of_quantity'])
                })
                niveles.append(3)
        
        # Agregar línea separadora después de cada brand
        filas.append({
            'Brand': '─' * 50,
            'Count of Case Number': '',
            'Sum of Quantity': ''
        })
        niveles.append(-1)

    # Agregar Grand Total al final
    filas.append({
        'Brand': 'GRAND TOTAL',
        'Count of Case Number': int(grand_total_cases),
        'Sum of Quantity': int(grand_total_qty)
    })
    niveles.append(0)

    tabla_jerarquica = pd.DataFrame(filas)

    # Escribir la tabla
    tabla_jerarquica.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

    # Dar formato de tabla de Excel
    wb = writer.book
    ws = wb[sheet_name]

    # Estilos
    font_brand = Font(bold=True, size=12, color="1F4E79")
    font_reason = Font(bold=True, size=11, color="2E75B6")
    font_detail = Font(size=10, color="404040")
    font_total = Font(bold=True, size=12, color="FFFFFF")
    font_separator = Font(color="CCCCCC")
    
    fill_brand = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    fill_reason = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    fill_detail = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_total = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_separator = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
    
    border_bottom = Border(bottom=Side(style='thin', color='CCCCCC'))
    border_thick = Border(bottom=Side(style='medium', color='1F4E79'))

    # Aplicar formato según el nivel
    for i, nivel in enumerate(niveles):
        row_num = startrow + 2 + i  # +2 porque startrow es 0-based y hay encabezado
        
        for col in range(1, 4):  # Columnas A, B, C
            cell = ws.cell(row=row_num, column=col)
            
            if nivel == 1:  # Brand
                cell.font = font_brand
                cell.fill = fill_brand
                cell.border = border_bottom
            elif nivel == 2:  # Reason
                cell.font = font_reason
                cell.fill = fill_reason
            elif nivel == 3:  # Detail
                cell.font = font_detail
                cell.fill = fill_detail
            elif nivel == 0:  # Grand Total
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_thick
            elif nivel == -1:  # Separador
                cell.font = font_separator
                cell.fill = fill_separator

    # Formato del encabezado
    for col in range(1, 4):
        cell = ws.cell(row=startrow + 1, column=col)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    print(f"Hoja '{sheet_name}': {len(tabla_jerarquica)} filas en tabla 2.")
    return startrow + len(tabla_jerarquica) + 2