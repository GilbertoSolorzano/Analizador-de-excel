import os
import json
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import google.generativeai as genai

# Cargar variables del .env
load_dotenv()

# Configurar Google Gemini
genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))
model = genai.GenerativeModel('gemini-2.0-flash')


def _sintetizar_descriptions_gemini(descriptions_rows: list, serie: str, reason: str, detail_reason: str) -> list:
    """
    Usa Google Gemini para agrupar y sintetizar las descripciones en categorías.
    Recibe: [{"description": str, "cas_qty": int, "qty_rejected": int}, ...]
    Devuelve: [{"defect": str, "cas_qty": int, "qty_rejected": int}, ...]
    """
    try:
        # Limitar para no enviar demasiado
        top = sorted(descriptions_rows, key=lambda x: x["qty_rejected"], reverse=True)[:60]

        prompt = f"""
Vas a agrupar y sintetizar descripciones de fallas para un reporte de Excel.

Contexto:
- Serie: {serie}
- Reason: {reason}
- Sub-detail (detail reason): {detail_reason}

Datos (cada item ya trae sus totales):
{json.dumps(top, ensure_ascii=False)}

Tarea:
1) Agrupa items con significado similar en 5 a 8 categorías (nombre corto en inglés, estilo: "Reader / Fob Failure", "Power & Battery Issues").
2) Suma cas_qty y qty_rejected por categoría.
3) Ordena por qty_rejected en orden descendente.

IMPORTANTE: Responde SOLO con un JSON array con este formato exacto, sin texto adicional ni explicaciones:
[
  {{"defect":"nombre categoria", "cas_qty":123, "qty_rejected":456}},
  {{"defect":"otra categoria", "cas_qty":78, "qty_rejected":90}}
]
"""

        response = model.generate_content(prompt)
        text = response.text.strip()

        # Limpiar si viene con markdown
        if "```" in text:
            # Extraer contenido entre ```
            parts = text.split("```")
            if len(parts) >= 2:
                text = parts[1]
                # Quitar "json" si está al inicio
                if text.startswith("json"):
                    text = text[4:]
                text = text.strip()

        # Parsear JSON
        data = json.loads(text)

        # Validar y formatear salida
        out = []
        for r in data:
            out.append({
                "defect": str(r.get("defect", ""))[:60],
                "cas_qty": int(r.get("cas_qty", 0)),
                "qty_rejected": int(r.get("qty_rejected", 0)),
            })
        return out

    except Exception as e:
        print(f"Error en Gemini: {e}")
        # Si falla, devolver los datos originales (top 10)
        return [
            {"defect": it["description"][:60] if it["description"] else "Sin descripción", 
             "cas_qty": it["cas_qty"], 
             "qty_rejected": it["qty_rejected"]}
            for it in descriptions_rows[:10]
        ]


def tabla_5(
    df_filtrado: pd.DataFrame,
    writer,
    sheet_name: str,
    col_serie,
    col_reason,
    col_detail_reason,
    col_description,
    col_case,
    col_qty,
    startrow=0,
    startcol=5,
    usar_ia=True
):
    """
    Genera tablas horizontales con:
    - Fila 1: Serie (verde)
    - Fila 2: Reason (azul)
    - Fila 3: Detail Reason (azul oscuro)
    - Fila 4: Encabezados (azul claro)
    - Filas de datos: Defect | CAS Qty | Qty Rejected
    
    Si usar_ia=True, sintetiza las descripciones con Google Gemini.
    """
    # Preparar resumen
    resumen = df_filtrado.loc[:, [col_serie, col_reason, col_detail_reason, col_description, col_case, col_qty]].copy()
    resumen.columns = ['serie', 'reason', 'detail_reason', 'description', 'case_number', 'quantity']
    resumen['quantity'] = pd.to_numeric(resumen['quantity'], errors='coerce').fillna(0)

    wb = writer.book
    ws = wb[sheet_name]

    # Obtener combinaciones únicas de serie + reason + detail_reason
    totales_grupo = (
        resumen
        .groupby(['serie', 'reason', 'detail_reason'], dropna=False, as_index=False)
        .agg(total_qty=('quantity', 'sum'))
        .sort_values('total_qty', ascending=False)
    )

    # Estilos
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    fill_serie = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Verde
    fill_reason = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # Azul
    fill_detail_reason = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")  # Azul oscuro
    fill_header = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Azul claro

    font_bold = Font(bold=True, size=11)
    font_header = Font(bold=True, size=10)
    center = Alignment(horizontal='center')

    tabla_count = 0
    current_col = startcol
    max_row_used = startrow

    for _, grupo_row in totales_grupo.iterrows():
        serie = grupo_row['serie']
        reason = grupo_row['reason']
        detail_reason = grupo_row['detail_reason']

        # Agregar por description
        datos = (
            resumen[
                (resumen['serie'] == serie) &
                (resumen['reason'] == reason) &
                (resumen['detail_reason'] == detail_reason)
            ]
            .groupby('description', dropna=False, as_index=False)
            .agg(
                cas_qty=('case_number', 'nunique'),
                qty_rejected=('quantity', 'sum')
            )
            .sort_values('qty_rejected', ascending=False)
        )

        if datos.empty:
            continue

        # Preparar items
        items = [
            {
                "description": ("" if pd.isna(r["description"]) else str(r["description"])),
                "cas_qty": int(r["cas_qty"]),
                "qty_rejected": int(r["qty_rejected"])
            }
            for _, r in datos.iterrows()
        ]

        # Sintetizar con IA o usar datos directos
        if usar_ia:
            try:
                filas_finales = _sintetizar_descriptions_gemini(items, serie, reason, detail_reason)
            except Exception as e:
                print(f"Tabla5 IA falló ({serie} / {reason} / {detail_reason}): {e}")
                filas_finales = [
                    {"defect": it["description"][:60] if it["description"] else "Sin descripción", 
                     "cas_qty": it["cas_qty"], 
                     "qty_rejected": it["qty_rejected"]}
                    for it in items[:15]
                ]
        else:
            filas_finales = [
                {"defect": it["description"][:60] if it["description"] else "Sin descripción", 
                 "cas_qty": it["cas_qty"], 
                 "qty_rejected": it["qty_rejected"]}
                for it in items
            ]

        # Escribir tabla
        tabla_count += 1
        current_row = startrow
        col_start = current_col
        col_end = current_col + 2

        # === FILA 1: Serie (Verde) ===
        ws.merge_cells(start_row=current_row + 1, start_column=col_start, end_row=current_row + 1, end_column=col_end)
        c = ws.cell(row=current_row + 1, column=col_start, value=str(serie))
        c.font = font_bold
        c.fill = fill_serie
        c.alignment = center
        c.border = border_thin
        for col in range(col_start, col_end + 1):
            ws.cell(row=current_row + 1, column=col).border = border_thin
        current_row += 1

        # === FILA 2: Reason (Azul) ===
        ws.merge_cells(start_row=current_row + 1, start_column=col_start, end_row=current_row + 1, end_column=col_end)
        c = ws.cell(row=current_row + 1, column=col_start, value=str(reason))
        c.font = font_bold
        c.fill = fill_reason
        c.alignment = center
        c.border = border_thin
        for col in range(col_start, col_end + 1):
            ws.cell(row=current_row + 1, column=col).border = border_thin
        current_row += 1

        # === FILA 3: Detail Reason (Azul oscuro) ===
        ws.merge_cells(start_row=current_row + 1, start_column=col_start, end_row=current_row + 1, end_column=col_end)
        c = ws.cell(row=current_row + 1, column=col_start, value=str(detail_reason))
        c.font = font_bold
        c.fill = fill_detail_reason
        c.alignment = center
        c.border = border_thin
        for col in range(col_start, col_end + 1):
            ws.cell(row=current_row + 1, column=col).border = border_thin
        current_row += 1

        # === FILA 4: Encabezados (Azul claro) ===
        headers = ['Defect', 'CAS Qty', 'Qty Rejected']
        for i, header in enumerate(headers):
            cell = ws.cell(row=current_row + 1, column=col_start + i, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = center
            cell.border = border_thin
        current_row += 1

        # === FILAS DE DATOS ===
        for r in filas_finales:
            cell_defect = ws.cell(row=current_row + 1, column=col_start, value=r["defect"])
            cell_defect.border = border_thin

            cell_cas = ws.cell(row=current_row + 1, column=col_start + 1, value=int(r["cas_qty"]))
            cell_cas.alignment = center
            cell_cas.border = border_thin

            cell_qty = ws.cell(row=current_row + 1, column=col_start + 2, value=int(r["qty_rejected"]))
            cell_qty.alignment = center
            cell_qty.border = border_thin

            current_row += 1

        # Actualizar fila máxima usada
        if current_row > max_row_used:
            max_row_used = current_row

        # Siguiente tabla a la derecha (3 columnas + 1 espacio)
        current_col += 4

    print(f"Hoja '{sheet_name}': {tabla_count} tablas 5 creadas (horizontal) {'con IA' if usar_ia else 'sin IA'}.")
    return max_row_used + 2