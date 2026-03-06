import pandas as pd 
import os
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
_table_counter = [0]  
def pedir_archivo():
    while True:
        nombre = input("Ingrese el nombre del archivo Excel (con extensión .xlsx o .xls): ").strip()
        if not (nombre.lower().endswith(".xlsx") or nombre.lower().endswith(".xls")):
            print("El archivo debe tener extensión .xlsx o .xls.")
            continue
        if not os.path.isfile(nombre):
            print(f" El archivo '{nombre}' no existe en la carpeta actual")
            continue
        return nombre
def pedir_hoja(archivo):
    try:
        xls = pd.ExcelFile(archivo)
        hojas = xls.sheet_names
    except Exception as e:
        print(f"No se pudo leer el archivo: {e}")
        return None

    hojas_lower = [h.lower() for h in hojas]
    while True:
        nombre_hoja = input("Ingresa el nombre de la hoja: ").strip()
        if nombre_hoja.lower() in hojas_lower:
            return hojas[hojas_lower.index(nombre_hoja.lower())]
        print(f"La hoja '{nombre_hoja}' no existe. Hojas disponibles: {', '.join(hojas)}")   
def leer_archivo(archivo, hoja): 
    try:
        datos = pd.read_excel(archivo, sheet_name=hoja, header=6, engine='openpyxl') 
        return datos
    except FileNotFoundError as e:
        print(f"Error: archivo no encontrado: {e}")
    except ValueError as e:
        print(f"Error leyendo la hoja: {e}")
    except Exception as e:
        print(f"Ocurrió un error inesperado: {e}")
    return None   
def build_save_path(original_path: str, suffix: str = '_filtrado', out_ext: str = '.xlsx') -> str:
    p = Path(original_path)
    stem = p.stem  # nombre sin extensión
    parent = p.parent
    new_name = f"{stem}{suffix}{out_ext}"
    full = parent / new_name
    i = 1
    # Si ya existe, agrega un contador: nombre_filtrado(1).xlsx, etc.
    while full.exists():
        full = parent / f"{stem}{suffix}({i}){out_ext}"
        i += 1
    return str(full)
def match_column_by_keywords(df, keywords):
        cols = list(df.columns)
        cols_lower = [c.strip().lower() for c in cols]
        for kw in keywords:
            for i, c in enumerate(cols_lower):
                if kw.lower() in c:
                    return cols[i]
        return None
def _make_table(df, startrow, suffix):
    if df.empty:
        return
    _table_counter[0] += 1
    nrows, ncols = df.shape
    ref = f"A{startrow + 1}:{get_column_letter(ncols)}{startrow + nrows + 1}"
    table_name = f"Tabla_{_table_counter[0]}_{suffix}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tbl)
def autofit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if column_letter is None:
                column_letter = cell.column_letter
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        if column_letter:
            ws.column_dimensions[column_letter].width = max_length + 2
#generar tabla 1
def tabla_1(df_filtrado: pd.DataFrame, writer, sheet_name: str, col_serie, col_case, col_qty):
    # Preparar resumen
    resumen = df_filtrado.loc[:, [col_serie, col_case, col_qty]].copy()
    resumen.columns = ['serie', 'case_of_numer', 'quality']
    resumen['quality'] = pd.to_numeric(resumen['quality'], errors='coerce').fillna(0)

    # Tabla detalle por serie + case (solo para calcular, no se escribe)
    tabla_resumen = (
        resumen
        .groupby(['serie', 'case_of_numer'], dropna=False, as_index=False)
        .agg({'quality': 'sum'})
        .rename(columns={'quality': 'sum_of_quality'})
    )

    # Tabla agrupada por serie
    tabla_por_serie = (
        tabla_resumen
        .groupby('serie', as_index=False)
        .agg(
            cantidad_de_casos=('case_of_numer', 'nunique'),
            suma_de_las_cantidades=('sum_of_quality', 'sum')
        )
    )

    # Escribir solo la tabla por serie
    tabla_por_serie.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

    # Dar formato de tabla de Excel
    wb = writer.book
    ws = wb[sheet_name]

    def _make_table(df, startrow, suffix):
        if df.empty:
            return
        _table_counter[0] += 1
        nrows, ncols = df.shape
        ref = f"A{startrow + 1}:{get_column_letter(ncols)}{startrow + nrows + 1}"
        table_name = f"Tabla_{_table_counter[0]}_{suffix}"
        tbl = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tbl)

    _make_table(tabla_por_serie, 0, "por_serie")

    print(f"Hoja '{sheet_name}': {len(tabla_por_serie)} series escritas.")
    return len(tabla_por_serie) + 2  # retorna la siguiente fila disponible
#tabla 2
def tabla_2(df_filtrado: pd.DataFrame, writer, sheet_name: str, col_serie, col_case, col_qty, col_reason, col_detail_reason, startrow=0):
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # Preparar resumen
    resumen = df_filtrado.loc[:, [col_serie, col_reason, col_detail_reason, col_case, col_qty]].copy()
    resumen.columns = ['brand', 'reason', 'detail_reason', 'case_number', 'quantity']
    resumen['quantity'] = pd.to_numeric(resumen['quantity'], errors='coerce').fillna(0)

    # Totales por brand (nivel 1)
    totales_brand = (
        resumen
        .groupby('brand', dropna=False, as_index=False)
        .agg(
            count_of_case_number=('case_number', 'nunique'),
            sum_of_quantity=('quantity', 'sum')
        )
        .sort_values('sum_of_quantity', ascending=False)
    )

    # Totales por brand + reason (nivel 2)
    totales_reason = (
        resumen
        .groupby(['brand', 'reason'], dropna=False, as_index=False)
        .agg(
            count_of_case_number=('case_number', 'nunique'),
            sum_of_quantity=('quantity', 'sum')
        )
    )

    # Detalle por brand + reason + detail_reason (nivel 3)
    detalle = (
        resumen
        .groupby(['brand', 'reason', 'detail_reason'], dropna=False, as_index=False)
        .agg(
            count_of_case_number=('case_number', 'nunique'),
            sum_of_quantity=('quantity', 'sum')
        )
    )

    # Grand Total
    grand_total_cases = resumen['case_number'].nunique()
    grand_total_qty = resumen['quantity'].sum()

    # Construir tabla jerárquica con niveles
    filas = []
    niveles = []  # Para guardar el nivel de cada fila (1=brand, 2=reason, 3=detail, 0=total, -1=separador)
    
    for idx, row_brand in totales_brand.iterrows():
        brand = row_brand['brand']
        
        # Nivel 1: Brand
        filas.append({
            'Brand': f"+ {brand}",
            'Count of Case Number': int(row_brand['count_of_case_number']),
            'Sum of Quantity': int(row_brand['sum_of_quantity'])
        })
        niveles.append(1)
        
        # Nivel 2: Reasons de este brand
        reasons_de_brand = totales_reason[totales_reason['brand'] == brand].sort_values('sum_of_quantity', ascending=False)
        for _, row_reason in reasons_de_brand.iterrows():
            reason = row_reason['reason']
            filas.append({
                'Brand': f"    - {reason}",
                'Count of Case Number': int(row_reason['count_of_case_number']),
                'Sum of Quantity': int(row_reason['sum_of_quantity'])
            })
            niveles.append(2)
            
            # Nivel 3: Details de esta reason
            detalles_de_reason = detalle[(detalle['brand'] == brand) & (detalle['reason'] == reason)].sort_values('sum_of_quantity', ascending=False)
            for _, det in detalles_de_reason.iterrows():
                filas.append({
                    'Brand': f"          • {det['detail_reason']}",
                    'Count of Case Number': int(det['count_of_case_number']),
                    'Sum of Quantity': int(det['sum_of_quantity'])
                })
                niveles.append(3)
        
        # Agregar línea separadora después de cada brand
        filas.append({
            'Brand': '─' * 50,
            'Count of Case Number': '',
            'Sum of Quantity': ''
        })
        niveles.append(-1)

    # Agregar Grand Total al final
    filas.append({
        'Brand': 'GRAND TOTAL',
        'Count of Case Number': int(grand_total_cases),
        'Sum of Quantity': int(grand_total_qty)
    })
    niveles.append(0)

    tabla_jerarquica = pd.DataFrame(filas)

    # Escribir la tabla
    tabla_jerarquica.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

    # Dar formato de tabla de Excel
    wb = writer.book
    ws = wb[sheet_name]

    # Estilos
    font_brand = Font(bold=True, size=12, color="1F4E79")
    font_reason = Font(bold=True, size=11, color="2E75B6")
    font_detail = Font(size=10, color="404040")
    font_total = Font(bold=True, size=12, color="FFFFFF")
    font_separator = Font(color="CCCCCC")
    
    fill_brand = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    fill_reason = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    fill_detail = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_total = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_separator = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
    
    border_bottom = Border(bottom=Side(style='thin', color='CCCCCC'))
    border_thick = Border(bottom=Side(style='medium', color='1F4E79'))

    # Aplicar formato según el nivel
    for i, nivel in enumerate(niveles):
        row_num = startrow + 2 + i  # +2 porque startrow es 0-based y hay encabezado
        
        for col in range(1, 4):  # Columnas A, B, C
            cell = ws.cell(row=row_num, column=col)
            
            if nivel == 1:  # Brand
                cell.font = font_brand
                cell.fill = fill_brand
                cell.border = border_bottom
            elif nivel == 2:  # Reason
                cell.font = font_reason
                cell.fill = fill_reason
            elif nivel == 3:  # Detail
                cell.font = font_detail
                cell.fill = fill_detail
            elif nivel == 0:  # Grand Total
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_thick
            elif nivel == -1:  # Separador
                cell.font = font_separator
                cell.fill = fill_separator

    # Formato del encabezado
    for col in range(1, 4):
        cell = ws.cell(row=startrow + 1, column=col)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    print(f"Hoja '{sheet_name}': {len(tabla_jerarquica)} filas en tabla 2.")
    return startrow + len(tabla_jerarquica) + 2

#genera tabla 3
def tabla_3(df_filtrado: pd.DataFrame, writer, sheet_name: str, col_case, col_qty, col_reason, col_detail_reason, startrow=0):
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # Preparar resumen
    resumen = df_filtrado.loc[:, [col_reason, col_detail_reason, col_case, col_qty]].copy()
    resumen.columns = ['reason', 'detail_reason', 'case_number', 'quantity']
    resumen['quantity'] = pd.to_numeric(resumen['quantity'], errors='coerce').fillna(0)

    # Totales por reason (nivel 1)
    totales_reason = (
        resumen
        .groupby('reason', dropna=False, as_index=False)
        .agg(
            count_of_case_number=('case_number', 'nunique'),
            sum_of_quantity=('quantity', 'sum')
        )
        .sort_values('sum_of_quantity', ascending=False)
    )

    # Detalle por reason + detail_reason (nivel 2)
    detalle = (
        resumen
        .groupby(['reason', 'detail_reason'], dropna=False, as_index=False)
        .agg(
            count_of_case_number=('case_number', 'nunique'),
            sum_of_quantity=('quantity', 'sum')
        )
    )

    # Grand Total
    grand_total_cases = resumen['case_number'].nunique()
    grand_total_qty = resumen['quantity'].sum()

    # Construir tabla jerárquica
    filas = []
    niveles = []
    
    for _, row_reason in totales_reason.iterrows():
        reason = row_reason['reason']
        
        # Nivel 1: Reason
        filas.append({
            'Reason': f"+{reason}",
            'Count of Case Number': int(row_reason['count_of_case_number']),
            'Sum of Quantity': int(row_reason['sum_of_quantity'])
        })
        niveles.append(1)
        
        # Nivel 2: Details de esta reason
        detalles_de_reason = detalle[detalle['reason'] == reason].sort_values('sum_of_quantity', ascending=False)
        for _, det in detalles_de_reason.iterrows():
            filas.append({
                'Reason': f"      • {det['detail_reason']}",
                'Count of Case Number': int(det['count_of_case_number']),
                'Sum of Quantity': int(det['sum_of_quantity'])
            })
            niveles.append(2)
        
        # Línea separadora después de cada reason
        filas.append({
            'Reason': '─' * 50,
            'Count of Case Number': '',
            'Sum of Quantity': ''
        })
        niveles.append(-1)

    # Grand Total
    filas.append({
        'Reason': 'GRAND TOTAL',
        'Count of Case Number': int(grand_total_cases),
        'Sum of Quantity': int(grand_total_qty)
    })
    niveles.append(0)

    tabla_jerarquica = pd.DataFrame(filas)

    # Escribir la tabla
    tabla_jerarquica.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

    # Dar formato
    wb = writer.book
    ws = wb[sheet_name]

    # Estilos
    font_reason = Font(bold=True, size=12, color="1F4E79")
    font_detail = Font(size=10, color="404040")
    font_total = Font(bold=True, size=12, color="FFFFFF")
    font_separator = Font(color="CCCCCC")
    
    fill_reason = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    fill_detail = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_total = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_separator = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
    
    border_bottom = Border(bottom=Side(style='thin', color='CCCCCC'))
    border_thick = Border(bottom=Side(style='medium', color='1F4E79'))

    # Aplicar formato según el nivel
    for i, nivel in enumerate(niveles):
        row_num = startrow + 2 + i
        
        for col in range(1, 4):
            cell = ws.cell(row=row_num, column=col)
            
            if nivel == 1:  # Reason
                cell.font = font_reason
                cell.fill = fill_reason
                cell.border = border_bottom
            elif nivel == 2:  # Detail
                cell.font = font_detail
                cell.fill = fill_detail
            elif nivel == 0:  # Grand Total
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_thick
            elif nivel == -1:  # Separador
                cell.font = font_separator
                cell.fill = fill_separator

    # Formato del encabezado
    for col in range(1, 4):
        cell = ws.cell(row=startrow + 1, column=col)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    print(f"Hoja '{sheet_name}': {len(tabla_jerarquica)} filas en tabla 3.")
    return startrow + len(tabla_jerarquica) + 2

## se genera tabla 4
def tabla_4(df_filtrado: pd.DataFrame, writer, sheet_name: str, col_customer, col_case, startrow=0):
    # Preparar resumen solo con customer y case
    resumen = df_filtrado.loc[:, [col_customer, col_case]].copy()
    resumen.columns = ['customer', 'case_of_number']

    tabla_por_customer = (
        resumen
        .groupby('customer', as_index=False)
        .agg(total_de_casos=('case_of_number', 'nunique'))
    )

    # Escribir a partir de startrow (no en 0)
    tabla_por_customer.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

    wb = writer.book
    ws = wb[sheet_name]

    def _make_table(df, startrow, suffix):
        if df.empty:
            return
        _table_counter[0] += 1
        nrows, ncols = df.shape
        ref = f"A{startrow + 1}:{get_column_letter(ncols)}{startrow + nrows + 1}"
        table_name = f"Tabla_{_table_counter[0]}_{suffix}"
        tbl = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tbl)

    _make_table(tabla_por_customer, startrow, "por_customer")
    print(f"Hoja '{sheet_name}': {len(tabla_por_customer)} clientes escritos.")
def guardar_filtros_en_hojas(datos: pd.DataFrame,  original_path: str):
    save_path = build_save_path(original_path, suffix='_filtrado', out_ext='.xlsx')
    #Aplica varios filtros y guarda cada resultado en una hoja distinta del mismo archivo Excel.
    
    # Define tus filtros: (nombre_hoja, columna, lista_de_valores)
    filtros = [
        #('Factory_Ensenada_Sauzal_Olathe', 'Factory', ['Ensenada', 'El Sauzal', 'Olathe']),
        ('Schlage_Residential_Mechanical', 'Brand / Category', ['Schlage Residential Mechanical']),
        ('Schlage_Residential_Electronic', 'Brand / Category', ['Schlage Residential Electronic']),
        ('Schlage_Electronic_Locks', 'Brand / Category', ['Schlage Electronic Locks']),
        ('Falcon_Lock', 'Brand / Category', ['Falcon - Lock']),
        ('Schlage_Commercial', 'Brand / Category', ['Schlage Commercial'])
    ]
    ###############

    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        for nombre, col, valores in filtros:
            sheet_name = nombre
            # Comprueba que columna exista
            if col not in datos.columns:
                print(f"Advertencia: la columna '{col}' no existe en el DataFrame. Hoja '{sheet_name}' vacía.")
                # opcional: f un DataFrame vacío o con aviso
                pd.DataFrame({'Aviso': [f"Columna '{col}' no encontrada"]}).to_excel(writer, sheet_name=sheet_name, index=False)
                continue

            # Filtrado (maneja valores nulos sin error)
            mask = datos[col].isin(valores)
            df_filtrado = datos[mask].copy()

            if df_filtrado.empty:
                print(f"No se encontraron filas para {nombre}. Se escribirá hoja vacía con mensaje.")
                pd.DataFrame({'Aviso': [f"No se encontraron filas para filtro: {valores}"]}).to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # Intentar detectar columnas para el resumen: serie, case, cantidad
                col_serie = match_column_by_keywords(df_filtrado, ['serie', 'serial', 'serial number', 's/n'])
                col_case  = match_column_by_keywords(df_filtrado, ['case', 'case of', 'case number', 'case_of', 'case#'])
                col_qty   = match_column_by_keywords(df_filtrado, ['quality', 'qty', 'quantity', 'quiality', 'cant', 'count'])
                col_customer = match_column_by_keywords(df_filtrado, ['customer', 'cliente', 'client', 'cust'])
                col_reason = match_column_by_keywords(df_filtrado, ['reason (english)', 'razon', 'razón'])
                col_detail_reason = match_column_by_keywords(df_filtrado, ['detail reason (english)', 'detalle', 'detail'])

                if col_serie is None or col_case is None or col_qty is None:
                    # Si falta alguna columna, avisamos y escribimos el df completo como antes
                    print(f"Advertencia: no se detectaron las columnas necesarias para el resumen: serie={col_serie}, case={col_case}, qty={col_qty}")
                    df_filtrado.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    next_row = tabla_1(df_filtrado, writer, sheet_name, col_serie, col_case, col_qty)

                    if col_reason is not None and col_detail_reason is not None:
                        next_row = tabla_2(df_filtrado, writer, sheet_name, col_serie, col_case, col_qty, col_reason, col_detail_reason, startrow=next_row)

                    
                    if col_reason is not None and col_detail_reason is not None:
                        next_row = tabla_3(df_filtrado, writer, sheet_name, col_case, col_qty, col_reason, col_detail_reason, startrow=next_row)
                    if col_customer is not None:
                        tabla_4(df_filtrado, writer, sheet_name, col_customer, col_case, startrow=next_row)
                        
                    ws = writer.book[sheet_name]
                    autofit_columns(ws)

if __name__ == "__main__": 
    archivo = 'ipl feb.xlsx'
    hoja = 'IPL - Cases'
    print(f"archivo seleccionado: {archivo}")
    print(f"hoja seleccionada: {hoja}")
    datos = leer_archivo(archivo, hoja) 

    if datos is None:
        print("No se pudo leer el DataFrame. Saliendo.")
    else:
        guardar_filtros_en_hojas(datos, original_path=archivo)
        